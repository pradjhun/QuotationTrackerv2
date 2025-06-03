import pandas as pd
from typing import Tuple

def validate_excel_structure(df: pd.DataFrame) -> Tuple[bool, str]:
    """
    Validate that the Excel file has the expected structure.
    
    Args:
        df: DataFrame to validate
        
    Returns:
        Tuple of (is_valid: bool, message: str)
    """
    if df.empty:
        return False, "The Excel file is empty."
    
    # Clean column names for comparison
    columns = [col.strip().upper() for col in df.columns]
    
    # Expected columns (some variations allowed)
    expected_columns = [
        'SL.NO', 'MODULE', 'BODY COLOUR', 'PICTURE', 
        'SINGLE COLOUR OPTION', 'WATT', 'SIZE', 'BEAM ANGLE', 'CUT OUT'
    ]
    
    # Alternative column names that are acceptable
    column_alternatives = {
        'SL.NO': ['SERIAL NO', 'SERIAL NUMBER', 'S.NO', 'SNO'],
        'BODY COLOUR': ['BODY COLOR'],
        'SINGLE COLOUR OPTION': ['SINGLE COLOR OPTION', 'SINGLE COLOR'],
        'CUT OUT': ['CUTOUT', 'CUT-OUT']
    }
    
    # Check if we have at least some key columns
    key_columns = ['MODULE', 'WATT', 'SIZE']
    found_key_columns = []
    
    for key_col in key_columns:
        if key_col in columns:
            found_key_columns.append(key_col)
        else:
            # Check alternatives
            alternatives = column_alternatives.get(key_col, [])
            for alt in alternatives:
                if alt in columns:
                    found_key_columns.append(key_col)
                    break
    
    if len(found_key_columns) < 2:
        return False, f"Missing key columns. Expected at least 2 of: {key_columns}. Found: {found_key_columns}"
    
    # Check for minimum number of rows
    if len(df) < 1:
        return False, "The Excel file must contain at least one data row."
    
    # Validate data types for numeric columns (allow mixed content)
    numeric_columns = ['WATT', 'SIZE']
    for col in numeric_columns:
        if col in df.columns:
            # Check if column has some valid numeric values or is mostly empty
            non_empty_values = df[col].dropna()
            if len(non_empty_values) > 0:
                numeric_series = pd.to_numeric(non_empty_values, errors='coerce')
                valid_numeric_ratio = numeric_series.notna().sum() / len(non_empty_values)
                # Allow if at least 30% of non-empty values are numeric or if all are empty/text
                if valid_numeric_ratio < 0.3 and len(non_empty_values) > 3:
                    return False, f"Column '{col}' should contain mostly numeric values. Found {valid_numeric_ratio:.1%} valid numbers."
    
    return True, f"File structure validated successfully. Found {len(df)} rows with {len(df.columns)} columns."

def format_dataframe_display(df: pd.DataFrame) -> pd.DataFrame:
    """
    Format DataFrame for better display in Streamlit.
    
    Args:
        df: DataFrame to format
        
    Returns:
        Formatted DataFrame
    """
    if df.empty:
        return df
    
    formatted_df = df.copy()
    
    # Format numeric columns
    numeric_columns = ['SL.NO', 'WATT', 'SIZE']
    for col in numeric_columns:
        if col in formatted_df.columns:
            # Convert to numeric and format
            numeric_series = pd.to_numeric(formatted_df[col], errors='coerce')
            # Format integers without decimal places
            formatted_df[col] = numeric_series.apply(
                lambda x: f"{int(x)}" if pd.notna(x) and x == int(x) else f"{x:.1f}" if pd.notna(x) else ""
            )
    
    # Ensure string columns are properly formatted
    string_columns = ['MODULE', 'BODY COLOUR', 'PICTURE', 'PRICE', 'BEAM ANGLE', 'CUT OUT']
    for col in string_columns:
        if col in formatted_df.columns:
            formatted_df[col] = formatted_df[col].astype(str).replace('nan', '').replace('None', '').fillna('')
    
    # Reorder columns to match expected order
    preferred_order = [
        'SL.NO', 'MODULE', 'BODY COLOUR', 'PICTURE', 
        'PRICE', 'WATT', 'SIZE', 'BEAM ANGLE', 'CUT OUT'
    ]
    
    # Only include columns that exist in the DataFrame
    ordered_columns = [col for col in preferred_order if col in formatted_df.columns]
    remaining_columns = [col for col in formatted_df.columns if col not in ordered_columns]
    final_order = ordered_columns + remaining_columns
    
    return formatted_df[final_order]

def export_to_excel(df: pd.DataFrame, filename: str = None, customer_name: str = "", customer_address: str = "", quotation_date: str = "", quotation_id: str = "", sales_person: str = "", sales_contact: str = "") -> bytes:
    """
    Export DataFrame to Excel format as bytes with embedded images matching the professional quotation format.
    
    Args:
        df: DataFrame to export
        filename: Optional filename (not used in bytes export)
        
    Returns:
        Excel file as bytes
    """
    import io
    import os
    from openpyxl import Workbook
    from openpyxl.drawing import image
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    # Create a copy of the dataframe and remove unnecessary columns
    df_export = df.copy()
    columns_to_remove = ['product_id', 'quotation_id', 'id']
    for col in columns_to_remove:
        if col in df_export.columns:
            df_export = df_export.drop(col, axis=1)
    
    # Add sequential serial number starting from 1
    df_export.insert(0, 'Product No', list(range(1, len(df_export) + 1)))
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    
    # Define styles
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    dark_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    
    current_row = 1
    
    # Title row - QUOTATION
    ws.merge_cells(f'A{current_row}:M{current_row}')
    title_cell = ws.cell(row=current_row, column=1, value="QUOTATION")
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = header_fill
    title_cell.border = border
    ws.row_dimensions[current_row].height = 30
    current_row += 1
    
    # Left section headers and Power Udyog section
    # Organization Name
    org_name_cell = ws.cell(row=current_row, column=1, value="Organization Name:")
    org_name_cell.font = Font(bold=True, size=10)
    org_name_cell.border = border
    ws.merge_cells(f'B{current_row}:F{current_row}')
    org_value_cell = ws.cell(row=current_row, column=2, value=customer_name or "")
    org_value_cell.border = border
    # Apply borders to merged cells
    for col in range(2, 7):
        ws.cell(row=current_row, column=col).border = border
    # Empty column G
    ws.cell(row=current_row, column=7).border = border
    
    # Store the row numbers for Power Udyog merging
    power_udyog_start_row = current_row
    
    # Apply borders to Power Udyog area for row 1
    for col in range(8, 14):
        ws.cell(row=current_row, column=col).border = border
    current_row += 1
    
    # Address
    addr_label_cell = ws.cell(row=current_row, column=1, value="Address:")
    addr_label_cell.font = Font(bold=True, size=10)
    addr_label_cell.border = border
    ws.merge_cells(f'B{current_row}:F{current_row}')
    addr_cell = ws.cell(row=current_row, column=2, value=customer_address or "")
    addr_cell.border = border
    addr_cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[current_row].height = 30
    # Apply borders to merged cells
    for col in range(2, 7):
        ws.cell(row=current_row, column=col).border = border
    # Empty column G
    ws.cell(row=current_row, column=7).border = border
    
    # Apply borders to Power Udyog area for row 2
    for col in range(8, 14):
        ws.cell(row=current_row, column=col).border = border
    
    # Merge column G from row 2 to row 6 (G2:G6)
    ws.merge_cells(f'G{power_udyog_start_row}:G{current_row + 3}')
    
    # Now merge Power Udyog across both rows (vertically and horizontally)
    ws.merge_cells(f'H{power_udyog_start_row}:M{current_row}')
    power_udyog_cell = ws.cell(row=power_udyog_start_row, column=8, value="Power Udyog")
    power_udyog_cell.font = Font(bold=True, size=36, color="1F4E79")
    power_udyog_cell.alignment = Alignment(horizontal='center', vertical='center')
    power_udyog_cell.border = border
    
    current_row += 1
    
    # Date
    date_cell = ws.cell(row=current_row, column=1, value="Date:")
    date_cell.font = Font(bold=True, size=10)
    date_cell.border = border
    
    # Merge cells B to F for date value and align left
    ws.merge_cells(f'B{current_row}:F{current_row}')
    date_value_cell = ws.cell(row=current_row, column=2, value=quotation_date or "")
    date_value_cell.border = border
    date_value_cell.alignment = Alignment(horizontal='left')
    # Apply borders to merged cells
    for col in range(2, 7):
        ws.cell(row=current_row, column=col).border = border
    
    # Quotation ID (right side)
    quot_id_cell = ws.cell(row=current_row, column=8, value="Quotation ID:")
    quot_id_cell.font = Font(bold=True, size=10)
    quot_id_cell.border = border
    
    # Merge cells I to M for quotation ID value and align left
    ws.merge_cells(f'I{current_row}:M{current_row}')
    quot_value_cell = ws.cell(row=current_row, column=9, value=quotation_id or "")
    quot_value_cell.border = border
    quot_value_cell.alignment = Alignment(horizontal='left')
    # Apply borders to merged cells
    for col in range(9, 14):
        ws.cell(row=current_row, column=col).border = border
    
    # Fill remaining cells with borders
    for col in [7]:
        ws.cell(row=current_row, column=col).border = border
    current_row += 1
    
    # PAN No and Sales Person
    pan_cell = ws.cell(row=current_row, column=1, value="PAN No")
    pan_cell.font = Font(bold=True, size=10)
    pan_cell.border = border
    
    # Merge cells B to F for PAN No value and align left
    ws.merge_cells(f'B{current_row}:F{current_row}')
    pan_value_cell = ws.cell(row=current_row, column=2, value="")
    pan_value_cell.border = border
    pan_value_cell.alignment = Alignment(horizontal='left')
    # Apply borders to merged cells
    for col in range(2, 7):
        ws.cell(row=current_row, column=col).border = border
    
    sales_cell = ws.cell(row=current_row, column=8, value="Sales Person:")
    sales_cell.font = Font(bold=True, size=10)
    sales_cell.border = border
    
    # Merge cells I to M for sales person value and align left
    ws.merge_cells(f'I{current_row}:M{current_row}')
    sales_value_cell = ws.cell(row=current_row, column=9, value=sales_person or "")
    sales_value_cell.border = border
    sales_value_cell.alignment = Alignment(horizontal='left')
    # Apply borders to merged cells
    for col in range(9, 14):
        ws.cell(row=current_row, column=col).border = border
    
    # Fill remaining cells with borders
    for col in [7]:
        ws.cell(row=current_row, column=col).border = border
    current_row += 1
    
    # GSTN No and Sales Contact
    gstn_cell = ws.cell(row=current_row, column=1, value="GSTN No")
    gstn_cell.font = Font(bold=True, size=10)
    gstn_cell.border = border
    
    # Merge cells B to F for GSTN No value and align left
    ws.merge_cells(f'B{current_row}:F{current_row}')
    gstn_value_cell = ws.cell(row=current_row, column=2, value="")
    gstn_value_cell.border = border
    gstn_value_cell.alignment = Alignment(horizontal='left')
    # Apply borders to merged cells
    for col in range(2, 7):
        ws.cell(row=current_row, column=col).border = border
    
    contact_cell = ws.cell(row=current_row, column=8, value="Sales Contact:")
    contact_cell.font = Font(bold=True, size=10)
    contact_cell.border = border
    
    # Merge cells I to M for sales contact value and align left
    ws.merge_cells(f'I{current_row}:M{current_row}')
    contact_value_cell = ws.cell(row=current_row, column=9, value=sales_contact or "")
    contact_value_cell.border = border
    contact_value_cell.alignment = Alignment(horizontal='left')
    # Apply borders to merged cells
    for col in range(9, 14):
        ws.cell(row=current_row, column=col).border = border
    
    # Fill remaining cells with borders
    for col in [7]:
        ws.cell(row=current_row, column=col).border = border
    current_row += 1
    
    # Product table headers
    headers = ['Product No', 'Model', 'Body Color', 'Product Image', 'Price', 'Watt', 'Size', 'Beam Angle', 'Cut Out', 'Light Color', 'Quantity', 'Discount', 'Item Total']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num, value=header)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = border
    current_row += 1
    
    # Product data rows with images
    for idx, row in df_export.iterrows():
        row_height = 60  # Default height for rows with images
        
        for col_idx, (col_name, value) in enumerate(row.items(), 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if col_name.lower() == 'picture' and value and os.path.exists(value):
                try:
                    # Add image to cell
                    img = image.Image(value)
                    img.width = 50  # Set image width
                    img.height = 50  # Set image height
                    ws.add_image(img, cell.coordinate)
                    cell.value = ""  # Don't add text value for image cell
                except:
                    cell.value = "Image not found"
            else:
                # Format numeric values appropriately
                if col_name.lower() in ['price', 'item_total']:
                    try:
                        cell.value = f"₹{float(value):,.2f}"
                    except:
                        cell.value = str(value)
                elif col_name.lower() == 'discount':
                    try:
                        cell.value = f"{float(value)}%"
                    except:
                        cell.value = str(value)
                else:
                    cell.value = str(value) if value is not None else ""
        
        ws.row_dimensions[current_row].height = row_height
        current_row += 1
    
    # Calculate totals
    subtotal = df_export['item_total'].sum() if 'item_total' in df_export.columns else 0
    gst_rate = 18.0
    gst_amount = subtotal * (gst_rate / 100)
    grand_total = subtotal + gst_amount
    
    # Add total rows
    current_row += 1  # Add spacing
    
    # Subtotal
    ws.merge_cells(f'L{current_row}:L{current_row}')
    subtotal_label = ws.cell(row=current_row, column=12, value="Subtotal:")
    subtotal_label.font = Font(bold=True)
    subtotal_label.alignment = Alignment(horizontal='right')
    subtotal_label.border = border
    
    subtotal_value = ws.cell(row=current_row, column=13, value=f"₹{subtotal:,.2f}")
    subtotal_value.font = Font(bold=True)
    subtotal_value.alignment = Alignment(horizontal='center')
    subtotal_value.border = border
    current_row += 1
    
    # GST
    gst_label = ws.cell(row=current_row, column=12, value=f"GST ({gst_rate}%):")
    gst_label.font = Font(bold=True)
    gst_label.alignment = Alignment(horizontal='right')
    gst_label.border = border
    
    gst_value = ws.cell(row=current_row, column=13, value=f"₹{gst_amount:,.2f}")
    gst_value.font = Font(bold=True)
    gst_value.alignment = Alignment(horizontal='center')
    gst_value.border = border
    current_row += 1
    
    # Grand Total
    grand_total_label = ws.cell(row=current_row, column=12, value="Grand Total:")
    grand_total_label.font = Font(bold=True)
    grand_total_label.alignment = Alignment(horizontal='right')
    grand_total_label.border = border
    
    grand_total_value = ws.cell(row=current_row, column=13, value=f"₹{grand_total:,.2f}")
    grand_total_value.font = Font(bold=True)
    grand_total_value.alignment = Alignment(horizontal='center')
    grand_total_value.border = border
    current_row += 2
    
    # Terms & Conditions and Bank Details sections
    terms_start_row = current_row
    
    # Terms & Conditions section
    ws.merge_cells(f'A{current_row}:F{current_row}')
    terms_header = ws.cell(row=current_row, column=1, value="TERMS & CONDITIONS")
    terms_header.font = Font(bold=True, size=12)
    terms_header.alignment = Alignment(horizontal='center')
    terms_header.fill = gray_fill
    terms_header.border = dark_border
    
    # Bank Details section
    bank_start_row = current_row
    ws.merge_cells(f'H{current_row}:M{current_row}')
    bank_header = ws.cell(row=current_row, column=8, value="Bank Details")
    bank_header.font = Font(bold=True, size=12)
    bank_header.alignment = Alignment(horizontal='center')
    bank_header.fill = gray_fill
    bank_header.border = border
    current_row += 1
    
    # Terms & Conditions content
    terms_conditions = [
        "GST & IGST ARE 18%",
        "100% ADVANCE PAYMENT",
        "PRICING ON FOB KOLKATA BASIS",
        "TWO YEAR WARRANTY ON LED",
        "TWO YEAR WARRANTY ON DRIVER",
        "SPOT LIGHTS ARE IP GRADED & DUSTPROOF",
        "DELIVERY WILL TAKE MINIMUM 10-35 WORKING DAYS FROM THE DATE OF CONFIRMED P.O AND ADVANCE PAYMENT.",
        "DELIVERY CHARGE EXTRA AS PER ACTUAL",
        "FOR EVERY BILLING GST NO OR PANCARD NO IS MANDATORY",
        "STRICTLY GOODS ONCE SOLD WILL NOT BE TAKEN BACK AS PER GST"
    ]
    
    # Bank details content
    bank_details = [
        ("Bank Name", "kotak Bank"),
        ("IFSC code", ""),
        ("Account No.", ""),
        ("Address", "")
    ]
    
    # Add terms conditions with no borders between rows
    terms_end_row = current_row + len(terms_conditions) - 1
    
    for i, term in enumerate(terms_conditions):
        if current_row + i <= terms_start_row + 10:  # Limit to available space
            # Merge cells A to F for each term
            ws.merge_cells(f'A{current_row + i}:F{current_row + i}')
            term_cell = ws.cell(row=current_row + i, column=1, value=term)
            term_cell.font = Font(size=9)
            term_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # No borders between rows - only apply outer borders to the section
            for col in range(1, 7):  # Columns A to F
                cell = ws.cell(row=current_row + i, column=col)
                # No internal borders, only outer border for the entire section
                cell.border = Border()
    
    # Add outer border around the entire Terms & Conditions section only
    for row in range(terms_start_row, terms_end_row + 2):  # Include header
        for col in range(1, 7):  # Columns A to F
            cell = ws.cell(row=row, column=col)
            
            # Apply outer border only to the edges of the section
            new_border = Border(
                left=Side(style='thin', color='000000') if col == 1 else None,
                right=Side(style='thin', color='000000') if col == 6 else None,
                top=Side(style='thin', color='000000') if row == terms_start_row else None,
                bottom=Side(style='thin', color='000000') if row == terms_end_row + 1 else None
            )
            cell.border = new_border
    
    # Bank Details content rows - matching the image format exactly
    # Bank Name row
    bank_name_cell = ws.cell(row=current_row, column=8, value="Bank Name")
    bank_name_cell.font = Font(size=10)
    bank_name_cell.border = border
    
    ws.merge_cells(f'I{current_row}:M{current_row}')
    bank_name_value = ws.cell(row=current_row, column=9, value="kotak Bank")
    bank_name_value.border = border
    # Apply border to all merged cells
    for col in range(9, 14):
        ws.cell(row=current_row, column=col).border = border
    current_row += 1
    
    # IFSC code row  
    ifsc_cell = ws.cell(row=current_row, column=8, value="IFSC code")
    ifsc_cell.font = Font(size=10)
    ifsc_cell.border = border
    
    ws.merge_cells(f'I{current_row}:M{current_row}')
    ifsc_value = ws.cell(row=current_row, column=9, value="")
    ifsc_value.border = border
    # Apply border to all merged cells
    for col in range(9, 14):
        ws.cell(row=current_row, column=col).border = border
    current_row += 1
    
    # Account No. row
    account_cell = ws.cell(row=current_row, column=8, value="Account No.")
    account_cell.font = Font(size=10)
    account_cell.border = border
    
    ws.merge_cells(f'I{current_row}:M{current_row}')
    account_value = ws.cell(row=current_row, column=9, value="")
    account_value.border = border
    # Apply border to all merged cells
    for col in range(9, 14):
        ws.cell(row=current_row, column=col).border = border
    current_row += 1
    
    # Address row
    address_cell = ws.cell(row=current_row, column=8, value="Address")
    address_cell.font = Font(size=10)
    address_cell.border = border
    
    ws.merge_cells(f'I{current_row}:M{current_row}')
    address_value = ws.cell(row=current_row, column=9, value="")
    address_value.border = border
    # Apply border to all merged cells
    for col in range(9, 14):
        ws.cell(row=current_row, column=col).border = border
    current_row += 1
    
    # Large Signature section - spanning to match Terms & Conditions height
    signature_start_row = current_row
    signature_rows_needed = len(terms_conditions) - 4  # Remaining rows to match terms height
    
    # Merge signature label area vertically to match terms height
    ws.merge_cells(f'H{current_row}:H{current_row + signature_rows_needed - 1}')
    signature_cell = ws.cell(row=current_row, column=8, value="Signature")
    signature_cell.font = Font(size=10)
    signature_cell.alignment = Alignment(horizontal='center', vertical='center')
    signature_cell.border = border
    
    # Merge signature value area to match terms height
    ws.merge_cells(f'I{current_row}:M{current_row + signature_rows_needed - 1}')
    signature_value = ws.cell(row=current_row, column=9, value="")
    signature_value.border = border
    
    # Apply borders to all cells in the signature area
    for sig_row in range(current_row, current_row + signature_rows_needed):
        for sig_col in range(8, 14):
            cell = ws.cell(row=sig_row, column=sig_col)
            cell.border = border
    
    # Set column widths for better appearance - Column A wider to match organization name
    column_widths = [20, 15, 12, 12, 12, 8, 15, 12, 12, 12, 10, 10, 12]  # Column A increased from 10 to 20
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def clean_search_term(search_term: str) -> str:
    """
    Clean and prepare search term for database queries.
    
    Args:
        search_term: Raw search term
        
    Returns:
        Cleaned search term
    """
    if not search_term:
        return ""
    
    # Remove extra whitespace and convert to lowercase
    cleaned = search_term.strip().lower()
    
    # Remove special characters that might interfere with search
    import re
    cleaned = re.sub(r'[^\w\s.-]', '', cleaned)
    
    return cleaned
